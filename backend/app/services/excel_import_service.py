# 엑셀 임포트 서비스
from datetime import date, time, datetime
from io import BytesIO
from uuid import UUID
import csv

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from app.db.models.salary import WorkRecord
from app.db.models.employee import Employee
from app.repositories.work_record_repo import WorkRecordRepository
from app.repositories.employee_repo import EmployeeRepository
from app.services.attendance_service import AttendanceService
from app.core.exceptions import ValidationError


class ParsedRow:
    """파싱된 엑셀 행"""
    def __init__(self, row_number: int, data: dict | None = None, errors: list[str] | None = None):
        self.row_number = row_number
        self.data = data
        self.errors = errors or []


class ExcelImportService:
    """엑셀/CSV 임포트 서비스"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.work_record_repo = WorkRecordRepository(db)
        self.employee_repo = EmployeeRepository(db)
        self.attendance_service = AttendanceService(db)

    async def parse_file(self, file: UploadFile) -> list[ParsedRow]:
        """
        엑셀/CSV 파일 파싱

        Returns:
            ParsedRow 리스트 (행 번호, 데이터, 에러 포함)
        """
        filename = file.filename or ""
        content = await file.read()

        if filename.endswith('.xlsx'):
            return await self._parse_xlsx(content)
        elif filename.endswith('.csv'):
            return await self._parse_csv(content)
        else:
            raise ValidationError(
                message="지원하지 않는 파일 형식입니다.",
                details=[{"field": "file", "message": "xlsx 또는 csv 파일을 업로드하세요."}]
            )

    async def _parse_xlsx(self, content: bytes) -> list[ParsedRow]:
        """엑셀 파일 파싱"""
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active

            parsed_rows = []
            header_row = None
            header_map = {}

            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                # 헤더 행 처리
                if row_idx == 1:
                    header_row = row
                    # 헤더 매핑
                    for col_idx, header in enumerate(header_row):
                        if header:
                            header_map[header.lower().strip()] = col_idx
                    continue

                # 빈 행 건너뛰기
                if not any(row):
                    continue

                # 데이터 행 파싱
                parsed_row = self._parse_xlsx_row(row_idx, row, header_map)
                parsed_rows.append(parsed_row)

            return parsed_rows

        except Exception as e:
            raise ValidationError(
                message="파일 파싱 실패",
                details=[{"field": "file", "message": str(e)}]
            )

    async def _parse_csv(self, content: bytes) -> list[ParsedRow]:
        """CSV 파일 파싱"""
        try:
            # UTF-8 BOM 제거
            if content.startswith(b'\xef\xbb\xbf'):
                content = content[3:]

            text = content.decode('utf-8')
            reader = csv.DictReader(text.splitlines())

            parsed_rows = []
            for row_idx, row in enumerate(reader, start=2):  # 헤더는 1행
                # 빈 행 건너뛰기
                if not any(row.values()):
                    continue

                parsed_row = self._parse_csv_row(row_idx, row)
                parsed_rows.append(parsed_row)

            return parsed_rows

        except Exception as e:
            raise ValidationError(
                message="파일 파싱 실패",
                details=[{"field": "file", "message": str(e)}]
            )

    def _parse_xlsx_row(self, row_number: int, row: tuple, header_map: dict) -> ParsedRow:
        """엑셀 행 파싱"""
        errors = []
        data = {}

        # 컬럼 매핑
        columns = {
            'employee_name': self._get_column_value(row, header_map, 'employee_name'),
            'work_date': self._get_column_value(row, header_map, 'work_date'),
            'scheduled_start': self._get_column_value(row, header_map, 'scheduled_start'),
            'scheduled_end': self._get_column_value(row, header_map, 'scheduled_end'),
            'actual_start': self._get_column_value(row, header_map, 'actual_start'),
            'actual_end': self._get_column_value(row, header_map, 'actual_end'),
            'break_minutes': self._get_column_value(row, header_map, 'break_minutes'),
            'is_holiday': self._get_column_value(row, header_map, 'is_holiday'),
            'memo': self._get_column_value(row, header_map, 'memo'),
        }

        # 필수 필드 확인
        if not columns['employee_name']:
            errors.append("필수 필드(employee_name) 누락")
        if not columns['work_date']:
            errors.append("필수 필드(work_date) 누락")
        if not columns['scheduled_start']:
            errors.append("필수 필드(scheduled_start) 누락")
        if not columns['scheduled_end']:
            errors.append("필수 필드(scheduled_end) 누락")

        if errors:
            return ParsedRow(row_number, None, errors)

        # 타입 변환
        try:
            data['employee_name'] = str(columns['employee_name']).strip()
            data['work_date'] = self._parse_date(columns['work_date'])
            data['scheduled_start'] = self._parse_time(columns['scheduled_start'])
            data['scheduled_end'] = self._parse_time(columns['scheduled_end'])
            data['actual_start'] = self._parse_time(columns['actual_start']) if columns['actual_start'] else None
            data['actual_end'] = self._parse_time(columns['actual_end']) if columns['actual_end'] else None
            data['break_minutes'] = int(columns['break_minutes']) if columns['break_minutes'] else 60
            data['is_holiday'] = self._parse_boolean(columns['is_holiday'])
            data['memo'] = str(columns['memo']).strip() if columns['memo'] else None

        except ValueError as e:
            errors.append(str(e))
            return ParsedRow(row_number, None, errors)

        # break_minutes 범위 검증
        if data['break_minutes'] < 0 or data['break_minutes'] > 480:
            errors.append("break_minutes는 0~480 범위만 허용됩니다.")

        if errors:
            return ParsedRow(row_number, None, errors)

        return ParsedRow(row_number, data)

    def _parse_csv_row(self, row_number: int, row: dict) -> ParsedRow:
        """CSV 행 파싱"""
        errors = []
        data = {}

        # 필수 필드 확인
        if not row.get('employee_name', '').strip():
            errors.append("필수 필드(employee_name) 누락")
        if not row.get('work_date', '').strip():
            errors.append("필수 필드(work_date) 누락")
        if not row.get('scheduled_start', '').strip():
            errors.append("필수 필드(scheduled_start) 누락")
        if not row.get('scheduled_end', '').strip():
            errors.append("필수 필드(scheduled_end) 누락")

        if errors:
            return ParsedRow(row_number, None, errors)

        # 타입 변환
        try:
            data['employee_name'] = str(row['employee_name']).strip()
            data['work_date'] = self._parse_date(row['work_date'])
            data['scheduled_start'] = self._parse_time(row['scheduled_start'])
            data['scheduled_end'] = self._parse_time(row['scheduled_end'])
            data['actual_start'] = self._parse_time(row.get('actual_start', '')) if row.get('actual_start', '').strip() else None
            data['actual_end'] = self._parse_time(row.get('actual_end', '')) if row.get('actual_end', '').strip() else None
            data['break_minutes'] = int(row.get('break_minutes', 60)) if row.get('break_minutes', '').strip() else 60
            data['is_holiday'] = self._parse_boolean(row.get('is_holiday', ''))
            data['memo'] = str(row.get('memo', '')).strip() if row.get('memo', '').strip() else None

        except ValueError as e:
            errors.append(str(e))
            return ParsedRow(row_number, None, errors)

        # break_minutes 범위 검증
        if data['break_minutes'] < 0 or data['break_minutes'] > 480:
            errors.append("break_minutes는 0~480 범위만 허용됩니다.")

        if errors:
            return ParsedRow(row_number, None, errors)

        return ParsedRow(row_number, data)

    def _get_column_value(self, row: tuple, header_map: dict, column_name: str):
        """열 값 추출"""
        col_idx = header_map.get(column_name.lower())
        if col_idx is not None and col_idx < len(row):
            return row[col_idx]
        return None

    def _parse_date(self, value: str | date) -> date:
        """날짜 파싱"""
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError(f"날짜 형식이 올바르지 않습니다. (YYYY-MM-DD): {value}")
        raise ValueError(f"날짜 형식이 올바르지 않습니다: {value}")

    def _parse_time(self, value: str | time) -> time:
        """시간 파싱"""
        if isinstance(value, time):
            return value
        if isinstance(value, str):
            try:
                return datetime.strptime(value, '%H:%M').time()
            except ValueError:
                raise ValueError(f"시간 형식이 올바르지 않습니다. (HH:MM): {value}")
        raise ValueError(f"시간 형식이 올바르지 않습니다: {value}")

    def _parse_boolean(self, value: str | bool | None) -> bool:
        """불린 파싱"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['예', 'y', '1', 'true', 'yes']
        return False

    async def validate_rows(
        self,
        company_id: UUID,
        parsed_rows: list[ParsedRow]
    ) -> tuple[list[ParsedRow], list[ParsedRow]]:
        """
        파싱된 행 검증

        Returns:
            (유효한 행, 오류 행)
        """
        valid_rows = []
        error_rows = []

        # 직원 목록 조회
        employees = await self.employee_repo.list_by_company(company_id, limit=1000)
        employee_map = {emp.name: emp for emp in employees}

        # 각 이름별로 중복 확인
        name_counts = {}
        for emp in employees:
            name_counts[emp.name] = name_counts.get(emp.name, 0) + 1

        seen_dates = {}  # (employee_id, work_date) 중복 확인

        for parsed_row in parsed_rows:
            if parsed_row.data is None:
                # 파싱 에러
                error_rows.append(parsed_row)
                continue

            errors = []
            data = parsed_row.data

            # 직원 존재 확인
            if data['employee_name'] not in employee_map:
                errors.append(f"직원을 찾을 수 없습니다. ({data['employee_name']})")
            elif name_counts.get(data['employee_name'], 0) > 1:
                errors.append(f"동일 이름의 직원이 여러 명입니다. employee_id를 사용해주세요.")
            else:
                # 파일 내 중복 확인
                employee_id = employee_map[data['employee_name']].id
                date_key = (employee_id, data['work_date'])

                if date_key in seen_dates:
                    errors.append("파일 내 중복 데이터입니다.")
                else:
                    seen_dates[date_key] = True

                    # DB 기존 데이터 중복 확인
                    existing = await self.work_record_repo.get_by_employee_and_date(
                        employee_id, data['work_date']
                    )
                    if existing:
                        errors.append("이미 근무 기록이 존재합니다.")

            if errors:
                parsed_row.errors.extend(errors)
                error_rows.append(parsed_row)
            else:
                # employee_name을 employee_id로 변환
                parsed_row.data['employee_id'] = employee_map[data['employee_name']].id
                valid_rows.append(parsed_row)

        return valid_rows, error_rows

    async def import_records(
        self,
        company_id: UUID,
        valid_rows: list[ParsedRow]
    ) -> dict:
        """유효한 행 일괄 임포트"""
        if not valid_rows:
            return {
                "total_rows": 0,
                "created": 0,
                "updated": 0,
                "skipped": 0,
                "errors": [],
            }

        created = 0
        updated = 0
        skipped = 0
        errors = []

        try:
            for parsed_row in valid_rows:
                data = parsed_row.data
                employee_id = data.pop('employee_id')

                try:
                    # 기존 기록 확인
                    existing = await self.work_record_repo.get_by_employee_and_date(
                        employee_id, data['work_date']
                    )

                    if existing:
                        # 업데이트
                        await self.attendance_service.update_record(
                            company_id=company_id,
                            record_id=existing.id,
                            **data
                        )
                        updated += 1
                    else:
                        # 생성
                        await self.attendance_service.create_record(
                            company_id=company_id,
                            employee_id=employee_id,
                            **data
                        )
                        created += 1

                except Exception as e:
                    errors.append({
                        "row": parsed_row.row_number,
                        "reason": str(e),
                    })
                    skipped += 1

            return {
                "total_rows": len(valid_rows) + len(errors),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }

        except Exception as e:
            raise ValidationError(
                message="임포트 중 오류 발생",
                details=[{"field": "import", "message": str(e)}]
            )

    async def generate_template(self, company_id: UUID) -> BytesIO:
        """엑셀 템플릿 생성"""
        workbook = Workbook()

        # 1. 템플릿 시트
        worksheet = workbook.active
        worksheet.title = "근무기록"

        # 헤더
        headers = [
            'employee_name',
            'work_date',
            'scheduled_start',
            'scheduled_end',
            'actual_start',
            'actual_end',
            'break_minutes',
            'is_holiday',
            'memo'
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 예시 데이터
        example_data = [
            ['홍길동', '2026-03-01', '09:00', '18:00', '08:55', '20:30', '60', '아니오', '프로젝트 마감'],
        ]

        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 컬럼 너비 조정
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 15
        worksheet.column_dimensions['I'].width = 20

        # 2. 직원 목록 시트
        employees = await self.employee_repo.list_by_company(company_id, limit=1000)

        emp_worksheet = workbook.create_sheet("직원목록")
        emp_headers = ['직원명', '직원ID']

        for col_idx, header in enumerate(emp_headers, start=1):
            cell = emp_worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, emp in enumerate(employees, start=2):
            emp_worksheet.cell(row=row_idx, column=1).value = emp.name
            emp_worksheet.cell(row=row_idx, column=2).value = str(emp.id)

        emp_worksheet.column_dimensions['A'].width = 15
        emp_worksheet.column_dimensions['B'].width = 40

        # 메모리에 저장
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer
